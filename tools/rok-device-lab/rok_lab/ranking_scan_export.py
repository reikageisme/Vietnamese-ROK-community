from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = (
    "rank",
    "name",
    "score",
    "scoreRaw",
    "evidenceImage",
    "needsReview",
)


def write_ranking_exports(
    directory: Path,
    records: list[dict[str, Any]],
    metadata: dict[str, Any],
    formats: set[str],
) -> dict[str, str]:
    directory.mkdir(parents=True, exist_ok=True)
    document = {"schemaVersion": 1, **metadata, "records": records}
    json_path = directory / "ranking.json"
    json_path.write_text(
        json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    paths = {"json": str(json_path.resolve())}

    if "jsonl" in formats:
        path = directory / "ranking.jsonl"
        path.write_text(
            "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in records),
            encoding="utf-8",
        )
        paths["jsonl"] = str(path.resolve())

    if "csv" in formats:
        path = directory / "ranking.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as output:
            writer = csv.DictWriter(output, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(records)
        paths["csv"] = str(path.resolve())

    if "xlsx" in formats:
        path = directory / "ranking.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = str(metadata.get("rankingType", "Ranking")).title()[:31]
        sheet.append(list(COLUMNS))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="173A5E")
        for row in records:
            sheet.append([row.get(column) for column in COLUMNS])
            evidence = row.get("evidenceImage")
            if evidence and Path(str(evidence)).is_file():
                image = ExcelImage(str(evidence))
                image.height = 28
                image.width = max(80, round(image.width * 28 / max(image.height, 1)))
                sheet.add_image(image, f"E{sheet.max_row}")
                sheet.row_dimensions[sheet.max_row].height = 26
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(COLUMNS, 1):
            values = [str(column), *(str(row.get(column, "")) for row in records)]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                44, max(12, max(len(value) for value in values) + 2)
            )
        workbook.save(path)
        paths["xlsx"] = str(path.resolve())
    return paths
