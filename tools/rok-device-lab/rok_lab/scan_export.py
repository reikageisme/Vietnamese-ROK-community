from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .governor import collector_record

COLUMNS = (
    "rank",
    "governorId",
    "name",
    "allianceTag",
    "allianceName",
    "power",
    "killPoints",
    "acclaim",
    "highestAcclaim",
    "deadTroops",
    "t1Kills",
    "t2Kills",
    "t3Kills",
    "t4Kills",
    "t5Kills",
    "totalKills",
    "t45Kills",
    "rangedPoints",
    "resourcesGathered",
    "resourceAssistance",
    "helps",
    "killsValidated",
    "tierKillPointsValidated",
    "needsReview",
)


def write_exports(
    directory: Path,
    records: list[dict[str, Any]],
    metadata: dict[str, Any],
    formats: set[str],
) -> dict[str, str]:
    directory.mkdir(parents=True, exist_ok=True)
    paths: dict[str, str] = {}
    collector = {
        "externalId": metadata["externalId"],
        "deviceId": metadata["deviceId"],
        "capturedAt": metadata["capturedAt"],
        "kingdom": {"number": metadata["kingdom"]},
        "coveragePercent": metadata["coveragePercent"],
        "evidenceObjectKeys": [],
        "records": [collector_record(record) for record in records],
    }
    json_path = directory / "scan.json"
    json_path.write_text(
        json.dumps(collector, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    paths["json"] = str(json_path.resolve())

    if "jsonl" in formats:
        path = directory / "governors.jsonl"
        path.write_text(
            "".join(
                json.dumps(record, ensure_ascii=False) + "\n" for record in records
            ),
            encoding="utf-8",
        )
        paths["jsonl"] = str(path.resolve())

    if "csv" in formats:
        path = directory / "governors.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as output:
            writer = csv.DictWriter(output, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(records)
        paths["csv"] = str(path.resolve())

    if "xlsx" in formats:
        path = directory / "governors.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"KD{metadata['kingdom']}"
        sheet.append(list(COLUMNS))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="173A5E")
        for record in records:
            sheet.append([record.get(column) for column in COLUMNS])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(COLUMNS, 1):
            values = [str(column), *(str(record.get(column, "")) for record in records)]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                42, max(10, max(len(value) for value in values) + 2)
            )
        workbook.save(path)
        paths["xlsx"] = str(path.resolve())
    return paths
